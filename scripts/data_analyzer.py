#!/usr/bin/env python3
"""
Data Analyzer Script
Analyzes all data files in the workspace to extract schemas, field names, and metadata
without loading entire large files into memory.
"""

import os
import csv
import sqlite3
import pandas as pd
import json
import re
from pathlib import Path
from typing import Dict, List, Any, Optional
import chardet
import zipfile
import openpyxl
from collections import defaultdict

class DataAnalyzer:
    def __init__(self, workspace_path: str = "."):
        self.workspace_path = Path(workspace_path)
        self.results = {
            "files": [],
            "summary": {
                "total_files": 0,
                "csv_files": 0,
                "sql_files": 0,
                "txt_files": 0,
                "xlsx_files": 0,
                "other_files": 0
            }
        }
        
    def detect_encoding(self, file_path: Path) -> str:
        """Detect file encoding"""
        try:
            with open(file_path, 'rb') as f:
                raw_data = f.read(10000)  # Read first 10KB
                result = chardet.detect(raw_data)
                return result['encoding'] or 'utf-8'
        except Exception:
            return 'utf-8'
    
    def analyze_csv_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze CSV file structure"""
        file_info = {
            "file_path": str(file_path),
            "file_type": "csv",
            "file_size_mb": file_path.stat().st_size / (1024 * 1024),
            "encoding": self.detect_encoding(file_path),
            "fields": [],
            "sample_data": [],
            "record_count": 0,
            "issues": []
        }
        
        try:
            encoding = file_info["encoding"]
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                # Read first few lines to get headers and sample data
                lines = []
                for i, line in enumerate(f):
                    if i >= 10:  # Read first 10 lines
                        break
                    lines.append(line.strip())
                
                if lines:
                    # Try to parse as CSV
                    try:
                        reader = csv.reader(lines)
                        headers = next(reader)
                        file_info["fields"] = headers
                        
                        # Get sample data
                        sample_count = 0
                        for row in reader:
                            if sample_count < 5:  # Get 5 sample rows
                                file_info["sample_data"].append(row)
                                sample_count += 1
                            else:
                                break
                        
                        # Estimate record count by counting lines
                        with open(file_path, 'r', encoding=encoding, errors='ignore') as f2:
                            file_info["record_count"] = sum(1 for _ in f2) - 1  # Subtract header
                            
                    except Exception as e:
                        file_info["issues"].append(f"CSV parsing error: {str(e)}")
                        # Try to extract fields from first line
                        if lines:
                            file_info["fields"] = lines[0].split(',')
                
        except Exception as e:
            file_info["issues"].append(f"File reading error: {str(e)}")
        
        return file_info
    
    def analyze_sql_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze SQL file structure"""
        file_info = {
            "file_path": str(file_path),
            "file_type": "sql",
            "file_size_mb": file_path.stat().st_size / (1024 * 1024),
            "encoding": self.detect_encoding(file_path),
            "tables": [],
            "sample_queries": [],
            "issues": []
        }
        
        try:
            encoding = file_info["encoding"]
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                content = f.read(10000)  # Read first 10KB
                
                # Extract table names
                table_pattern = r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?[`"]?(\w+)[`"]?\s*\('
                tables = re.findall(table_pattern, content, re.IGNORECASE)
                file_info["tables"] = list(set(tables))
                
                # Extract sample queries
                query_pattern = r'(SELECT\s+.*?;|INSERT\s+.*?;|UPDATE\s+.*?;|DELETE\s+.*?;)'
                queries = re.findall(query_pattern, content, re.IGNORECASE | re.DOTALL)
                file_info["sample_queries"] = queries[:5]  # First 5 queries
                
        except Exception as e:
            file_info["issues"].append(f"SQL file reading error: {str(e)}")
        
        return file_info
    
    def analyze_txt_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze TXT file structure"""
        file_info = {
            "file_path": str(file_path),
            "file_type": "txt",
            "file_size_mb": file_path.stat().st_size / (1024 * 1024),
            "encoding": self.detect_encoding(file_path),
            "line_count": 0,
            "sample_lines": [],
            "delimiter": None,
            "potential_fields": [],
            "issues": []
        }
        
        try:
            encoding = file_info["encoding"]
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                lines = []
                for i, line in enumerate(f):
                    if i >= 20:  # Read first 20 lines
                        break
                    lines.append(line.strip())
                
                file_info["line_count"] = len(lines)
                file_info["sample_lines"] = lines[:5]
                
                # Try to detect delimiter and potential fields
                if lines:
                    first_line = lines[0]
                    delimiters = [',', '\t', '|', ';', ' ']
                    
                    for delimiter in delimiters:
                        if delimiter in first_line:
                            parts = first_line.split(delimiter)
                            if len(parts) > 1:
                                file_info["delimiter"] = delimiter
                                file_info["potential_fields"] = [f"field_{i+1}" for i in range(len(parts))]
                                break
                
        except Exception as e:
            file_info["issues"].append(f"TXT file reading error: {str(e)}")
        
        return file_info
    
    def analyze_xlsx_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze XLSX file structure"""
        file_info = {
            "file_path": str(file_path),
            "file_type": "xlsx",
            "file_size_mb": file_path.stat().st_size / (1024 * 1024),
            "sheets": [],
            "sample_data": {},
            "issues": []
        }
        
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            file_info["sheets"] = workbook.sheetnames
            
            # Get sample data from first sheet
            if workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
                sample_data = []
                for row in sheet.iter_rows(max_row=5, values_only=True):
                    sample_data.append(row)
                file_info["sample_data"] = {workbook.sheetnames[0]: sample_data}
                
        except Exception as e:
            file_info["issues"].append(f"XLSX file reading error: {str(e)}")
        
        return file_info
    
    def analyze_file(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """Analyze a single file based on its extension"""
        if not file_path.is_file():
            return None
            
        file_info = None
        extension = file_path.suffix.lower()
        
        if extension == '.csv':
            file_info = self.analyze_csv_file(file_path)
        elif extension == '.sql':
            file_info = self.analyze_sql_file(file_path)
        elif extension == '.txt':
            file_info = self.analyze_txt_file(file_path)
        elif extension in ['.xlsx', '.xls']:
            file_info = self.analyze_xlsx_file(file_path)
        else:
            # Skip other file types for now
            return None
        
        if file_info:
            self.results["files"].append(file_info)
            self.results["summary"][f"{file_info['file_type']}_files"] += 1
            self.results["summary"]["total_files"] += 1
        
        return file_info
    
    def analyze_all_files(self):
        """Analyze all files in the workspace"""
        print("Starting data analysis...")
        
        for file_path in self.workspace_path.iterdir():
            if file_path.is_file():
                print(f"Analyzing: {file_path.name}")
                self.analyze_file(file_path)
        
        print(f"Analysis complete. Found {self.results['summary']['total_files']} data files.")
    
    def save_results(self, output_file: str = "data_analysis_results.json"):
        """Save analysis results to JSON file"""
        def json_serializer(obj):
            """Custom JSON serializer for objects not serializable by default json code"""
            if hasattr(obj, 'isoformat'):
                return obj.isoformat()
            else:
                return str(obj)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False, default=json_serializer)
        print(f"Results saved to {output_file}")
    
    def generate_summary_report(self) -> str:
        """Generate a summary report"""
        summary = []
        summary.append("# Data Analysis Summary Report")
        summary.append("")
        
        # File type summary
        summary.append("## File Type Distribution")
        for file_type, count in self.results["summary"].items():
            if file_type != "total_files":
                summary.append(f"- {file_type.replace('_', ' ').title()}: {count}")
        summary.append("")
        
        # Large files
        large_files = [f for f in self.results["files"] if f["file_size_mb"] > 10]
        if large_files:
            summary.append("## Large Files (>10MB)")
            for file_info in large_files:
                summary.append(f"- {Path(file_info['file_path']).name}: {file_info['file_size_mb']:.1f}MB")
            summary.append("")
        
        # Files with issues
        files_with_issues = [f for f in self.results["files"] if f.get("issues")]
        if files_with_issues:
            summary.append("## Files with Issues")
            for file_info in files_with_issues:
                summary.append(f"- {Path(file_info['file_path']).name}: {', '.join(file_info['issues'])}")
            summary.append("")
        
        return "\n".join(summary)

def main():
    """Main function"""
    analyzer = DataAnalyzer()
    analyzer.analyze_all_files()
    analyzer.save_results()
    
    # Generate and save summary report
    summary = analyzer.generate_summary_report()
    with open("data_analysis_summary.md", 'w', encoding='utf-8') as f:
        f.write(summary)
    
    print("Analysis complete! Check data_analysis_results.json and data_analysis_summary.md")

if __name__ == "__main__":
    main() 